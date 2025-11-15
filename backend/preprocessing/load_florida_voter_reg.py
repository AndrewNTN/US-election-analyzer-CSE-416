# python
import logging
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from pymongo import MongoClient

MONGO_URI = "mongodb://localhost:27017/"
DATABASE_NAME = "cse416"
COLLECTION_NAME = "state_voter_registration"

# Florida state FIPS code
FLORIDA_FIPS = "12"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _clean_numeric_value(value) -> Optional[int]:
    """Clean and convert numeric values from Excel, handling None and commas."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        # Remove commas and whitespace
        cleaned = value.strip().replace(',', '')
        if not cleaned or cleaned.lower() in ['', 'n/a', 'na']:
            return None
        try:
            return int(float(cleaned))
        except ValueError:
            logger.warning(f"Could not convert '{value}' to integer")
            return None
    return None


def _clean_county_name(value) -> str:
    """Clean county name from Excel."""
    if value is None:
        return ""
    return str(value).strip()


def load_florida_party_affiliation():
    """Load Florida party affiliation data from Excel file into MongoDB."""

    # Get the path to the Excel file
    script_dir = Path(__file__).parent
    excel_path = script_dir.parent / "src" / "main" / "resources" / "florida-party affiliation" / "voter-registration-report-archive-2024" / "party-affiliation-by-county-2024.xlsx"

    if not excel_path.exists():
        logger.error(f"File not found: {excel_path}")
        return

    logger.info(f"Reading Florida party affiliation data from {excel_path}...")

    # Load the Excel file
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        logger.info(f"Loaded worksheet: {sheet.title}")
    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        return

    # Parse the header row to find column indices
    # The Florida Excel file has a specific structure:
    # Row 1: Title
    # Row 2: Date info
    # Row 3: Section header
    # Row 4: Column headers
    # Row 5+: Data
    header_row = None
    data_start_row = None

    # Find header row (look for row with 'County' as first column)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] and str(row[0]).strip().lower() == 'county':
            header_row = row
            data_start_row = row_idx + 1
            logger.info(f"Found header row at row {row_idx}: {header_row}")
            break

    if header_row is None:
        logger.error("Could not find header row with 'County' column")
        workbook.close()
        return

    # Map column names to indices
    column_map = {}
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        col_name = str(cell_value).strip().lower()
        column_map[col_name] = idx

    logger.info(f"Found columns: {list(column_map.keys())}")

    # Identify the required columns based on actual Excel structure
    # Expected columns: County, Republican Party Of Florida, Florida Democratic Party, Minor Party, No Party Affiliation, Totals
    county_col = None
    total_col = None
    republican_col = None
    democrat_col = None
    unaffiliated_col = None

    # Search for county column
    for key in column_map:
        if 'county' == key:
            county_col = column_map[key]
            break

    # Search for total column
    for key in column_map:
        if 'total' in key:
            total_col = column_map[key]
            break

    # Search for party columns - match the actual column names from Florida data
    for key in column_map:
        if 'republican' in key:
            republican_col = column_map[key]
        elif 'democratic' in key or 'democrat' in key:
            democrat_col = column_map[key]
        elif 'no party' in key or 'unaffiliated' in key:
            unaffiliated_col = column_map[key]

    logger.info(f"Column mapping - County: {county_col}, Total: {total_col}, Republican: {republican_col}, Democrat: {democrat_col}, Unaffiliated: {unaffiliated_col}")

    if county_col is None:
        logger.error("Could not find County column")
        workbook.close()
        return

    if total_col is None:
        logger.error("Could not find Total column")
        workbook.close()
        return

    # Parse data rows
    county_data: List[Dict] = []
    state_totals = {
        'totalRegisteredVoters': 0,
        'democraticVoters': 0,
        'republicanVoters': 0,
        'unaffiliatedVoters': 0
    }

    rows_processed = 0
    rows_with_data = 0

    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        rows_processed += 1

        # Get county name
        county_name = _clean_county_name(row[county_col] if county_col < len(row) else None)

        # Skip empty or total rows
        if not county_name or county_name.lower() in ['total', 'totals', 'statewide', 'state total', '']:
            continue

        # Get numeric values
        total_voters = _clean_numeric_value(row[total_col] if total_col is not None and total_col < len(row) else None)
        republican_voters = _clean_numeric_value(row[republican_col] if republican_col is not None and republican_col < len(row) else None)
        democrat_voters = _clean_numeric_value(row[democrat_col] if democrat_col is not None and democrat_col < len(row) else None)
        unaffiliated_voters = _clean_numeric_value(row[unaffiliated_col] if unaffiliated_col is not None and unaffiliated_col < len(row) else None)

        # Data completeness check
        missing_fields = []
        if total_voters is None:
            missing_fields.append('total')
        if republican_voters is None:
            missing_fields.append('republican')
        if democrat_voters is None:
            missing_fields.append('democrat')
        if unaffiliated_voters is None:
            missing_fields.append('unaffiliated')

        if missing_fields:
            logger.warning(f"County '{county_name}' is missing data for: {', '.join(missing_fields)}")

        # Create county record (use 0 for missing values)
        county_record = {
            'countyName': county_name,
            'totalRegisteredVoters': total_voters if total_voters is not None else 0,
            'democraticVoters': democrat_voters if democrat_voters is not None else 0,
            'republicanVoters': republican_voters if republican_voters is not None else 0,
            'unaffiliatedVoters': unaffiliated_voters if unaffiliated_voters is not None else 0
        }

        county_data.append(county_record)
        rows_with_data += 1

        # Add to state totals
        state_totals['totalRegisteredVoters'] += county_record['totalRegisteredVoters']
        state_totals['democraticVoters'] += county_record['democraticVoters']
        state_totals['republicanVoters'] += county_record['republicanVoters']
        state_totals['unaffiliatedVoters'] += county_record['unaffiliatedVoters']

    workbook.close()

    logger.info(f"Processed {rows_processed} rows, found {rows_with_data} counties with data")

    # Data completeness summary
    logger.info(f"Data completeness check:")
    logger.info(f"  Total counties processed: {rows_with_data}")
    logger.info(f"  State totals:")
    logger.info(f"    Total Registered Voters: {state_totals['totalRegisteredVoters']:,}")
    logger.info(f"    Democratic Voters: {state_totals['democraticVoters']:,}")
    logger.info(f"    Republican Voters: {state_totals['republicanVoters']:,}")
    logger.info(f"    Unaffiliated Voters: {state_totals['unaffiliatedVoters']:,}")

    # Validate totals
    party_sum = state_totals['democraticVoters'] + state_totals['republicanVoters'] + state_totals['unaffiliatedVoters']
    if state_totals['totalRegisteredVoters'] > 0:
        coverage_pct = (party_sum / state_totals['totalRegisteredVoters']) * 100
        logger.info(f"  Party coverage: {coverage_pct:.2f}% of total registered voters")
        if coverage_pct < 95:
            logger.warning(f"Party coverage is low ({coverage_pct:.2f}%), there may be other party affiliations not captured")

    # Create state document matching StateVoterRegistration model
    state_document = {
        'stateFips': FLORIDA_FIPS,
        'totalRegisteredVoters': state_totals['totalRegisteredVoters'],
        'democraticVoters': state_totals['democraticVoters'],
        'republicanVoters': state_totals['republicanVoters'],
        'unaffiliatedVoters': state_totals['unaffiliatedVoters'],
        'countyVoterRegistrations': county_data
    }

    # Connect to MongoDB
    client = MongoClient(MONGO_URI)
    db = client[DATABASE_NAME]
    col = db[COLLECTION_NAME]

    # Delete existing Florida data
    deleted_count = col.delete_many({'stateFips': FLORIDA_FIPS}).deleted_count
    logger.info(f"Deleted {deleted_count} existing Florida records")

    # Insert the state document
    try:
        result = col.insert_one(state_document)
        logger.info(f"Successfully inserted Florida voter registration data with ID: {result.inserted_id}")
    except Exception as e:
        logger.exception(f"Failed to insert data: {e}")
        client.close()
        return

    # Create indexes
    try:
        col.create_index("stateFips", unique=True)
        logger.info("Created index on stateFips")
    except Exception as e:
        logger.warning(f"Index creation failed (may already exist): {e}")

    # Verify the data
    count = col.count_documents({'stateFips': FLORIDA_FIPS})
    logger.info(f"Collection now contains {count} Florida document(s)")

    # Show sample
    sample = col.find_one({'stateFips': FLORIDA_FIPS})
    if sample:
        logger.info(f"Sample document (first 3 counties):")
        logger.info(f"  State FIPS: {sample['stateFips']}")
        logger.info(f"  Total Registered: {sample['totalRegisteredVoters']:,}")
        logger.info(f"  Democratic: {sample['democraticVoters']:,}")
        logger.info(f"  Republican: {sample['republicanVoters']:,}")
        logger.info(f"  Unaffiliated: {sample['unaffiliatedVoters']:,}")
        logger.info(f"  Number of counties: {len(sample['countyVoterRegistrations'])}")
        for i, county in enumerate(sample['countyVoterRegistrations'][:3]):
            logger.info(f"    County {i+1}: {county}")

    client.close()
    logger.info("Load complete.")


if __name__ == "__main__":
    load_florida_party_affiliation()

